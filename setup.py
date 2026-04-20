"""
OrderPro — setup.py
====================
Run this ONCE before starting the app:
    python setup.py

What it does:
  1. Checks all required libraries
  2. Creates a professional, formatted orders.xlsx (3 sheets)
  3. Confirms templates folder exists
"""

import os, sys
from collections import defaultdict
from datetime import datetime

# ═══════════════════════════════════════════════════════════
#  STEP 1: CHECK LIBRARIES
# ═══════════════════════════════════════════════════════════
REQUIRED = {
    "flask":      "flask",
    "pandas":     "pandas",
    "openpyxl":   "openpyxl",
    "sklearn":    "scikit-learn",
    "matplotlib": "matplotlib",
}

print("\n╔══════════════════════════════════════════════════╗")
print("║         OrderPro v3.0 — Setup                   ║")
print("╚══════════════════════════════════════════════════╝\n")

print("📦 Checking libraries...")
missing = []
for mod, pkg in REQUIRED.items():
    try:
        __import__(mod)
        print(f"   ✅  {pkg}")
    except ImportError:
        print(f"   ❌  {pkg}  ← MISSING")
        missing.append(pkg)

# pywhatkit needs browser/GUI — just check if installed, don't import
try:
    import importlib.util
    spec = importlib.util.find_spec("pywhatkit")
    if spec:
        print(f"   ✅  pywhatkit")
    else:
        print(f"   ⚠️  pywhatkit  ← install karo: pip install pywhatkit")
except Exception:
    pass

if missing:
    print(f"\n⚠️  Missing libraries install karo:")
    print(f"   pip install {' '.join(missing)} pywhatkit\n")
    sys.exit(1)

# ═══════════════════════════════════════════════════════════
#  STEP 2: CREATE PROFESSIONAL orders.xlsx
# ═══════════════════════════════════════════════════════════
print("\n📊 Creating professional orders.xlsx...")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Demo Data (fictional — safe for public GitHub) ────────
ORDERS_DATA = [
    ("ORD-001", "Ahmed Raza",          "03001111001", "Premium Wireless Headphones",   3500.00, "Delivered", "2026-01-05"),
    ("ORD-002", "Sara Khan",           "03002222002", "Leather Laptop Bag",            2800.00, "Delivered", "2026-01-07"),
    ("ORD-003", "Bilal Chaudhry",      "03003333003", "Mechanical Keyboard",           4200.00, "Delivered", "2026-01-10"),
    ("ORD-004", "Fatima Malik",        "03004444004", "USB-C Hub 7-in-1",              1800.00, "Delivered", "2026-01-12"),
    ("ORD-005", "Usman Tariq",         "03005555005", "Smart Watch Series 8",          9500.00, "Delivered", "2026-01-15"),
    ("ORD-006", "Ayesha Siddiqui",     "03006666006", "Portable Bluetooth Speaker",   2200.00, "Delivered", "2026-01-18"),
    ("ORD-007", "Hassan Javed",        "03007777007", "Noise Cancelling Earbuds",      5500.00, "Delivered", "2026-01-20"),
    ("ORD-008", "Zainab Hussain",      "03008888008", "Webcam HD 1080p",               3100.00, "Delivered", "2026-01-22"),
    ("ORD-009", "Kamran Sheikh",       "03009999009", "Gaming Mouse RGB",              1950.00, "Delivered", "2026-01-25"),
    ("ORD-010", "Nadia Iqbal",         "03001234510", "Desk Organizer Set",             850.00, "Delivered", "2026-01-28"),
    ("ORD-011", "Ahmed Raza",          "03001111001", "Monitor Stand Adjustable",      2100.00, "Delivered", "2026-02-02"),
    ("ORD-012", "Sara Khan",           "03002222002", "Wireless Charging Pad",         1200.00, "Delivered", "2026-02-05"),
    ("ORD-013", "Bilal Chaudhry",      "03003333003", "Smart Home LED Bulbs (4 Pack)", 1600.00, "Delivered", "2026-02-08"),
    ("ORD-014", "Tariq Mehmood",       "03001234514", "Portable Power Bank 20000mAh",  2750.00, "Delivered", "2026-02-10"),
    ("ORD-015", "Hina Baig",           "03001234515", "Laptop Cooling Pad",            1450.00, "Delivered", "2026-02-13"),
    ("ORD-016", "Omer Farooq",         "03001234516", "Ergonomic Office Chair",       18500.00, "Delivered", "2026-02-15"),
    ("ORD-017", "Mariam Zahid",        "03001234517", "Standing Desk Converter",      12000.00, "Sent",      "2026-02-18"),
    ("ORD-018", "Asad Nawaz",          "03001234518", "4K Portable Monitor",          22000.00, "Sent",      "2026-02-20"),
    ("ORD-019", "Sana Mirza",          "03001234519", "Smart Security Camera",         7800.00, "Sent",      "2026-02-22"),
    ("ORD-020", "Junaid Qureshi",      "03001234520", "Mechanical Numpad",             2400.00, "Sent",      "2026-02-25"),
    ("ORD-021", "Ahmed Raza",          "03001111001", "NVMe SSD 1TB",                  8500.00, "Pending",   "2026-03-01"),
    ("ORD-022", "Fatima Malik",        "03004444004", "RAM 16GB DDR5",                 6200.00, "Pending",   "2026-03-03"),
    ("ORD-023", "Hassan Javed",        "03007777007", "GPU Cooling Fan Set",           3300.00, "Pending",   "2026-03-05"),
    ("ORD-024", "Nadia Iqbal",         "03001234510", "Cable Management Kit",           950.00, "Pending",   "2026-03-07"),
    ("ORD-025", "Kamran Sheikh",       "03009999009", "Dual Monitor Arm",              5800.00, "Pending",   "2026-03-09"),
    ("ORD-026", "Sana Mirza",          "03001234519", "Laptop Privacy Screen",         2600.00, "Pending",   "2026-03-10"),
    ("ORD-027", "Mariam Zahid",        "03001234517", "Smart Plug Wi-Fi (2 Pack)",     1800.00, "Pending",   "2026-03-12"),
    ("ORD-028", "Usman Tariq",         "03005555005", "Portable Projector Mini",      14500.00, "Pending",   "2026-03-14"),
    ("ORD-029", "Ayesha Siddiqui",     "03006666006", "Digital Drawing Tablet",        8900.00, "Pending",   "2026-03-15"),
    ("ORD-030", "Tariq Mehmood",       "03001234514", "VR Headset Standalone",        35000.00, "Pending",   "2026-03-16"),
]

# ── Color Palette ─────────────────────────────────────────
C_HDR_BG  = "1E1B4B"
C_HDR_FG  = "FFFFFF"
C_TITLE   = "4F46E5"
C_DEL     = "D1FAE5"
C_SENT    = "EDE9FE"
C_PEND    = "FEF3C7"
C_ALT     = "F8F7FF"
C_BORDER  = "CBD5E1"

thin = Side(style="thin", color=C_BORDER)
bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

STATUS_BG = {"Delivered": C_DEL,   "Sent": C_SENT,   "Pending": C_PEND}
STATUS_FG = {"Delivered": "065F46", "Sent": "5B21B6", "Pending": "92400E"}

wb = Workbook()

# ════════════════════════════════════════
#  SHEET 1 — Orders
# ════════════════════════════════════════
ws = wb.active
ws.title = "Orders"

# Title row
ws.merge_cells("A1:G1")
c = ws["A1"]
c.value     = "OrderPro — Orders Register"
c.font      = Font(name="Arial", size=15, bold=True, color=C_HDR_FG)
c.fill      = PatternFill("solid", fgColor=C_TITLE)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 36

# Subtitle row
ws.merge_cells("A2:G2")
c = ws["A2"]
c.value     = f"Generated: {datetime.now().strftime('%B %d, %Y')}  |  Demo Data — OrderPro v3.0"
c.font      = Font(name="Arial", size=9, italic=True, color="6B7280")
c.fill      = PatternFill("solid", fgColor="EEF2FF")
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 20

ws.append([])  # blank row 3

# Header row
HEADERS = ["Order ID", "Customer Name", "Phone", "Product / Description", "Amount (Rs)", "Status", "Date"]
ws.append(HEADERS)
for col in range(1, 8):
    c = ws.cell(row=4, column=col)
    c.font      = Font(name="Arial", size=10, bold=True, color=C_HDR_FG)
    c.fill      = PatternFill("solid", fgColor=C_HDR_BG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = bdr
ws.row_dimensions[4].height = 22

# Data rows
for i, row in enumerate(ORDERS_DATA, start=5):
    order_id, name, phone, product, amount, status, date = row
    ws.append([order_id, name, phone, product, amount, status, date])
    alt = C_ALT if i % 2 == 0 else "FFFFFF"

    for col in range(1, 8):
        c = ws.cell(row=i, column=col)
        c.font      = Font(name="Arial", size=10)
        c.border    = bdr
        c.alignment = Alignment(
            vertical="center",
            horizontal="right" if col == 5 else "center" if col in [1, 6, 7] else "left"
        )
        if col == 6:
            c.fill = PatternFill("solid", fgColor=STATUS_BG.get(status, "FFFFFF"))
            c.font = Font(name="Arial", size=10, bold=True, color=STATUS_FG.get(status, "000000"))
        else:
            c.fill = PatternFill("solid", fgColor=alt)

    ws.cell(row=i, column=5).number_format = "#,##0.00"
    ws.cell(row=i, column=7).number_format = "YYYY-MM-DD"
    ws.row_dimensions[i].height = 20

# Total row
sum_row = len(ORDERS_DATA) + 5
c = ws.cell(row=sum_row, column=4)
c.value     = "TOTAL"
c.font      = Font(name="Arial", size=10, bold=True, color=C_HDR_FG)
c.fill      = PatternFill("solid", fgColor=C_HDR_BG)
c.alignment = Alignment(horizontal="right", vertical="center")
c.border    = bdr

c = ws.cell(row=sum_row, column=5)
c.value         = f"=SUM(E5:E{sum_row - 1})"
c.font          = Font(name="Arial", size=11, bold=True, color="065F46")
c.fill          = PatternFill("solid", fgColor="D1FAE5")
c.number_format = "#,##0.00"
c.alignment     = Alignment(horizontal="right", vertical="center")
c.border        = bdr
ws.row_dimensions[sum_row].height = 24

for col in range(1, 8):
    ws.cell(row=sum_row, column=col).border = bdr

# Column widths & freeze
for i, width in enumerate([12, 24, 15, 34, 14, 12, 13], 1):
    ws.column_dimensions[get_column_letter(i)].width = width
ws.freeze_panes = "A5"

# ════════════════════════════════════════
#  SHEET 2 — Summary Dashboard
# ════════════════════════════════════════
ws2 = wb.create_sheet("Summary")

ws2.merge_cells("A1:D1")
c = ws2["A1"]
c.value     = "OrderPro — Summary Dashboard"
c.font      = Font(name="Arial", size=14, bold=True, color=C_HDR_FG)
c.fill      = PatternFill("solid", fgColor=C_TITLE)
c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 34

ws2.merge_cells("A2:D2")
c = ws2["A2"]
c.value     = f"Auto-calculated from Orders sheet  |  {datetime.now().strftime('%B %d, %Y')}"
c.font      = Font(name="Arial", size=9, italic=True, color="6B7280")
c.fill      = PatternFill("solid", fgColor="EEF2FF")
c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[2].height = 18

kpis = [
    ("A4",  "B4",  "Total Orders",     f"=COUNTA(Orders!A5:A{sum_row-1})",                    "FFFFFF", "1E1B4B"),
    ("A6",  "B6",  "Total Revenue",    f"=Orders!E{sum_row}",                                 "D1FAE5", "065F46"),
    ("A8",  "B8",  "Delivered Orders", f'=COUNTIF(Orders!F5:F{sum_row-1},"Delivered")',        "D1FAE5", "065F46"),
    ("A10", "B10", "Sent Orders",      f'=COUNTIF(Orders!F5:F{sum_row-1},"Sent")',             "EDE9FE", "5B21B6"),
    ("A12", "B12", "Pending Orders",   f'=COUNTIF(Orders!F5:F{sum_row-1},"Pending")',          "FEF3C7", "92400E"),
    ("A14", "B14", "Avg Order Value",  f"=Orders!E{sum_row}/COUNTA(Orders!A5:A{sum_row-1})",  "EEF2FF", "1E1B4B"),
]

for lc_addr, vc_addr, label, formula, bg, fg in kpis:
    lc = ws2[lc_addr]
    lc.value     = label
    lc.font      = Font(name="Arial", size=10, bold=True, color="374151")
    lc.fill      = PatternFill("solid", fgColor="F3F4F6")
    lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    lc.border    = bdr
    ws2.row_dimensions[int(lc_addr[1:])].height = 26

    vc = ws2[vc_addr]
    vc.value     = formula
    vc.font      = Font(name="Arial", size=13, bold=True, color=fg)
    vc.fill      = PatternFill("solid", fgColor=bg)
    vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.border    = bdr
    if "Revenue" in label or "Avg" in label:
        vc.number_format = '"Rs "#,##0.00'

for col, width in zip("ABCD", [22, 18, 5, 20]):
    ws2.column_dimensions[col].width = width

# ════════════════════════════════════════
#  SHEET 3 — Customers
# ════════════════════════════════════════
ws3 = wb.create_sheet("Customers")

ws3.merge_cells("A1:E1")
c = ws3["A1"]
c.value     = "Customer Register"
c.font      = Font(name="Arial", size=14, bold=True, color=C_HDR_FG)
c.fill      = PatternFill("solid", fgColor=C_TITLE)
c.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 34

ws3.append([])
ws3.append(["Customer Name", "Phone", "Total Orders", "Total Spent (Rs)", "Last Order Date"])
for col in range(1, 6):
    c = ws3.cell(row=3, column=col)
    c.font      = Font(name="Arial", size=10, bold=True, color=C_HDR_FG)
    c.fill      = PatternFill("solid", fgColor=C_HDR_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = bdr
ws3.row_dimensions[3].height = 22

cust = defaultdict(lambda: {"phone": "", "count": 0, "total": 0.0, "last": ""})
for row in ORDERS_DATA:
    n, ph, amt, dt = row[1], row[2], row[4], row[6]
    cust[n]["phone"] = ph
    cust[n]["count"] += 1
    cust[n]["total"] += amt
    cust[n]["last"]   = max(cust[n]["last"], dt)

for r, (name, info) in enumerate(sorted(cust.items()), start=4):
    ws3.append([name, info["phone"], info["count"], info["total"], info["last"]])
    alt = C_ALT if r % 2 == 0 else "FFFFFF"
    for col in range(1, 6):
        c = ws3.cell(row=r, column=col)
        c.font      = Font(name="Arial", size=10)
        c.fill      = PatternFill("solid", fgColor=alt)
        c.border    = bdr
        c.alignment = Alignment(
            vertical="center",
            horizontal="right" if col == 4 else "center" if col in [3, 5] else "left"
        )
    ws3.cell(row=r, column=4).number_format = "#,##0.00"
    ws3.row_dimensions[r].height = 20

for col, width in zip("ABCDE", [24, 16, 14, 18, 16]):
    ws3.column_dimensions[col].width = width
ws3.freeze_panes = "A4"

# ── Save workbook ─────────────────────
wb.save("orders.xlsx")
print("   ✅  orders.xlsx created — 3 professional sheets:")
print("       📋 Orders  |  📊 Summary  |  👥 Customers")

total_rev = sum(r[4] for r in ORDERS_DATA)
delivered = sum(1 for r in ORDERS_DATA if r[5] == "Delivered")
pending   = sum(1 for r in ORDERS_DATA if r[5] == "Pending")
sent      = sum(1 for r in ORDERS_DATA if r[5] == "Sent")
print(f"\n   📦 {len(ORDERS_DATA)} orders  |  💰 Rs {total_rev:,.2f} revenue")
print(f"   ✅ {delivered} Delivered  |  💬 {sent} Sent  |  ⏳ {pending} Pending")

# ═══════════════════════════════════════════════════════════
#  STEP 3: CHECK TEMPLATES FOLDER
# ═══════════════════════════════════════════════════════════
print("\n📁 Checking templates folder...")

templates_needed = [
    "base.html", "index.html", "orders.html", "customers.html",
    "reports.html", "pending.html", "whatsapp.html",
    "delivered.html", "settings.html", "login.html"
]

all_ok = True
for t in templates_needed:
    path = os.path.join("templates", t)
    if os.path.exists(path):
        print(f"   ✅  templates/{t}")
    else:
        print(f"   ❌  templates/{t}  ← MISSING")
        all_ok = False

print("\n" + "═" * 50)
if all_ok:
    print("🎉 Setup complete! Ab run karo:")
    print("   python app.py")
    print("\n🔐 Login credentials:")
    print("   admin    /  admin123")
    print("   manager  /  manager123")
else:
    print("⚠️  Kuch template files missing hain.")
    print("   Sab HTML files  →  templates\\  folder mein rakho.")
    print("   Phir run karo   →  python app.py")
print("═" * 50 + "\n")
