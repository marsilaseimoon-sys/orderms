"""
OrderPro v3.0 — Professional Flask App
=======================================
Run: python app.py
"""

from flask import Flask, render_template, request, jsonify, redirect, url_for, session, flash
import pandas as pd
import os, io, base64
from datetime import datetime
import webbrowser, threading
import pywhatkit

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

try:
    from sklearn.linear_model import LinearRegression
    import numpy as np
    AI_AVAILABLE = True
except ImportError:
    AI_AVAILABLE = False

app = Flask("OrderPro")
app.secret_key = "orderpro_secret_2026"

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, "orders.xlsx")

# ── Demo credentials (change as needed) ──────────────────
USERS = {
    "admin":   {"password": "admin123",  "role": "Admin",   "name": "Admin User"},
    "manager": {"password": "manager123","role": "Manager", "name": "Sales Manager"},
}

# ═══════════════════════════════════════════════════════════
#  AUTH HELPERS
# ═══════════════════════════════════════════════════════════
def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

# ═══════════════════════════════════════════════════════════
#  DATA HELPERS
# ═══════════════════════════════════════════════════════════
def load_orders():
    if not os.path.exists(EXCEL_FILE):
        return pd.DataFrame(columns=["OrderID","CustomerName","Phone","Product","Amount","Status","Date"])
    raw = pd.read_excel(EXCEL_FILE, sheet_name="Orders", header=None)
    header_row = 0
    for i, row in raw.iterrows():
        vals = [str(v).strip().lower() for v in row.values]
        if any("order" in v and "id" in v for v in vals):
            header_row = i; break
    df = pd.read_excel(EXCEL_FILE, sheet_name="Orders", header=header_row)
    col_map = {}
    for col in df.columns:
        c = str(col).strip().lower().replace(" ","").replace("/","").replace("(rs)","")
        if "orderid" in c:                           col_map[col] = "OrderID"
        elif "customername" in c:                    col_map[col] = "CustomerName"
        elif "phone" in c:                           col_map[col] = "Phone"
        elif "product" in c or "description" in c:  col_map[col] = "Product"
        elif "amount" in c:                          col_map[col] = "Amount"
        elif "status" in c:                          col_map[col] = "Status"
        elif "date" in c:                            col_map[col] = "Date"
    df = df.rename(columns=col_map)
    if "OrderID" in df.columns:
        df = df[df["OrderID"].notna() & df["OrderID"].astype(str).str.startswith("ORD")]
    for col in ["OrderID","CustomerName","Phone","Product","Amount","Status","Date"]:
        if col not in df.columns: df[col] = ""
    df["Amount"] = pd.to_numeric(df["Amount"], errors="coerce").fillna(0)
    df["Date"]   = pd.to_datetime(df["Date"], errors="coerce")
    return df.reset_index(drop=True)

def save_orders(df):
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    df2 = df.copy()
    df2["Date"] = df2["Date"].dt.strftime("%Y-%m-%d")
    C_HDR_BG="1E1B4B"; C_TITLE="4F46E5"
    C_DEL="D1FAE5"; C_SENT="EDE9FE"; C_PEND="FEF3C7"; C_ALT="F8F7FF"; C_BOR="CBD5E1"
    thin=Side(style='thin',color=C_BOR); bdr=Border(left=thin,right=thin,top=thin,bottom=thin)
    STATUS_BG={"Delivered":C_DEL,"Sent":C_SENT,"Pending":C_PEND}
    STATUS_FG={"Delivered":"065F46","Sent":"5B21B6","Pending":"92400E"}
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Orders"] if "Orders" in wb.sheetnames else wb.active
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
            for cell in row: cell.value = None
    else:
        wb = Workbook(); ws = wb.active; ws.title = "Orders"
    for i, (_, row) in enumerate(df2.iterrows(), start=5):
        vals = [row["OrderID"],row["CustomerName"],row["Phone"],row["Product"],row["Amount"],row["Status"],row["Date"]]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.font=Font(name='Arial',size=10); c.border=bdr
            alt = C_ALT if i%2==0 else "FFFFFF"
            c.alignment=Alignment(vertical='center',horizontal='right' if col==5 else 'center' if col in [1,6,7] else 'left')
            status = row["Status"]
            if col==6:
                c.fill=PatternFill('solid',fgColor=STATUS_BG.get(status,"FFFFFF"))
                c.font=Font(name='Arial',size=10,bold=True,color=STATUS_FG.get(status,"000000"))
            else:
                c.fill=PatternFill('solid',fgColor=alt)
        ws.cell(row=i,column=5).number_format='#,##0.00'
        ws.row_dimensions[i].height=20
    sr=len(df2)+5
    ws.cell(row=sr,column=4).value='TOTAL'
    ws.cell(row=sr,column=4).font=Font(name='Arial',size=10,bold=True,color="FFFFFF")
    ws.cell(row=sr,column=4).fill=PatternFill('solid',fgColor=C_HDR_BG)
    ws.cell(row=sr,column=4).alignment=Alignment(horizontal='right')
    ws.cell(row=sr,column=5).value=f'=SUM(E5:E{sr-1})'
    ws.cell(row=sr,column=5).font=Font(name='Arial',size=11,bold=True,color='065F46')
    ws.cell(row=sr,column=5).fill=PatternFill('solid',fgColor='D1FAE5')
    ws.cell(row=sr,column=5).number_format='#,##0.00'
    ws.cell(row=sr,column=5).alignment=Alignment(horizontal='right')
    for col in range(1,8): ws.cell(row=sr,column=col).border=bdr
    ws.row_dimensions[sr].height=24
    wb.save(EXCEL_FILE)

def fmt_phone(raw):
    raw = str(raw).strip()
    if raw.startswith("0"): return "+92"+raw[1:]
    if not raw.startswith("+"): return "+92"+raw
    return raw

def get_stats(df):
    return {
        "total":     len(df),
        "pending":   len(df[df["Status"]=="Pending"]),
        "sent":      len(df[df["Status"].isin(["Sent","Delivered"])]),
        "delivered": len(df[df["Status"]=="Delivered"]),
        "revenue":   f"Rs {df['Amount'].sum():,.0f}",
        "revenue_raw": float(df["Amount"].sum()),
    }

# ═══════════════════════════════════════════════════════════
#  CHARTS
# ═══════════════════════════════════════════════════════════
BG="#0d1117"; SURF="#161b22"; ACC="#6c63ff"; GRN="#00e5a0"; YLW="#ffd166"; RED="#ff6584"

def _b64(fig):
    buf=io.BytesIO(); fig.savefig(buf,format="png",bbox_inches="tight",facecolor=BG,dpi=110)
    buf.seek(0); plt.close(fig); return base64.b64encode(buf.read()).decode()

def chart_revenue(df):
    df2=df.dropna(subset=["Date"]).copy(); df2["Month"]=df2["Date"].dt.to_period("M")
    m=df2.groupby("Month")["Amount"].sum().reset_index(); m["Month"]=m["Month"].dt.to_timestamp()
    fig,ax=plt.subplots(figsize=(7,3)); fig.patch.set_facecolor(BG); ax.set_facecolor(SURF)
    ax.plot(m["Month"],m["Amount"],color=ACC,lw=2.5,marker="o",ms=6)
    ax.fill_between(m["Month"],m["Amount"],alpha=0.15,color=ACC)
    ax.set_title("Monthly Revenue Trend",color="#e6edf3",fontsize=11,pad=8)
    ax.tick_params(colors="#8b949e",labelsize=8); ax.spines[:].set_color("#30363d")
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %y"))
    return _b64(fig)

def chart_donut(df):
    counts=df["Status"].value_counts()
    colors=[ACC if l=="Sent" else GRN if l=="Delivered" else YLW if l=="Pending" else RED for l in counts.index]
    fig,ax=plt.subplots(figsize=(4,3)); fig.patch.set_facecolor(BG); ax.set_facecolor(BG)
    _,texts,autos=ax.pie(counts.values,labels=counts.index,colors=colors,autopct="%1.0f%%",startangle=90,wedgeprops=dict(width=0.55))
    for t in texts: t.set_color("#8b949e"); t.set_fontsize(9)
    for a in autos: a.set_color("white"); a.set_fontsize(9)
    ax.set_title("Order Status",color="#e6edf3",fontsize=11,pad=8)
    return _b64(fig)

def chart_products(df):
    top=df.groupby("Product")["Amount"].sum().nlargest(6).reset_index()
    fig,ax=plt.subplots(figsize=(6,3)); fig.patch.set_facecolor(BG); ax.set_facecolor(SURF)
    ax.barh(top["Product"],top["Amount"],color=GRN,height=0.55)
    ax.set_title("Top Products by Revenue",color="#e6edf3",fontsize=11,pad=8)
    ax.tick_params(colors="#8b949e",labelsize=8); ax.spines[:].set_color("#30363d"); ax.invert_yaxis()
    return _b64(fig)

def chart_customers(df):
    top=df.groupby("CustomerName")["Amount"].sum().nlargest(5).reset_index()
    fig,ax=plt.subplots(figsize=(6,3)); fig.patch.set_facecolor(BG); ax.set_facecolor(SURF)
    ax.bar(top["CustomerName"],top["Amount"],color=YLW,width=0.55)
    ax.set_title("Top Customers",color="#e6edf3",fontsize=11,pad=8)
    ax.tick_params(colors="#8b949e",labelsize=8); ax.spines[:].set_color("#30363d")
    plt.xticks(rotation=20,ha='right')
    return _b64(fig)

def ai_insights(df):
    if df.empty: return ["Add orders to get AI insights!"]
    ins=[]
    top_p=df.groupby("Product")["Amount"].sum().idxmax()
    ins.append(f"🏆 Best product: <strong>{top_p}</strong> — Rs {df[df['Product']==top_p]['Amount'].sum():,.0f}")
    top_c=df.groupby("CustomerName")["Amount"].sum().idxmax()
    ins.append(f"⭐ Top customer: <strong>{top_c}</strong> — Rs {df[df['CustomerName']==top_c]['Amount'].sum():,.0f} spent")
    pend=len(df[df["Status"]=="Pending"])
    if pend: ins.append(f"⚠️ <strong>{pend}</strong> pending orders need WhatsApp confirmation")
    if AI_AVAILABLE:
        try:
            df2=df.dropna(subset=["Date"]).copy()
            df2["M"]=df2["Date"].dt.to_period("M").apply(lambda p: p.ordinal)
            m=df2.groupby("M")["Amount"].sum().reset_index()
            if len(m)>=2:
                model=LinearRegression().fit(m[["M"]],m["Amount"])
                fc=model.predict([[m["M"].max()+1]])[0]
                ins.append(f"📈 AI Forecast: Next month ≈ <strong>Rs {max(fc,0):,.0f}</strong>")
        except: pass
    ins.append(f"💡 Average order value: <strong>Rs {df['Amount'].mean():,.0f}</strong>")
    rate=len(df[df["Status"]=="Delivered"])/len(df)*100
    ins.append(f"✅ Delivery rate: <strong>{rate:.0f}%</strong> ({len(df[df['Status']=='Delivered'])}/{len(df)})")
    return ins

# ═══════════════════════════════════════════════════════════
#  AUTH ROUTES
# ═══════════════════════════════════════════════════════════
@app.route("/login", methods=["GET","POST"])
def login():
    if "user" in session: return redirect(url_for("index"))
    error = None
    if request.method == "POST":
        username = request.form.get("username","").strip()
        password = request.form.get("password","").strip()
        if username in USERS and USERS[username]["password"] == password:
            session["user"]     = username
            session["role"]     = USERS[username]["role"]
            session["name"]     = USERS[username]["name"]
            return redirect(url_for("index"))
        error = "Invalid username or password!"
    return render_template("login.html", error=error)

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# ═══════════════════════════════════════════════════════════
#  MAIN ROUTES
# ═══════════════════════════════════════════════════════════
@app.route("/")
@login_required
def index():
    df=load_orders(); recent=df.tail(5).copy()
    recent["Date"]=recent["Date"].dt.strftime("%Y-%m-%d")
    return render_template("index.html",stats=get_stats(df),recent=recent.to_dict(orient="records"))

@app.route("/orders")
@login_required
def orders():
    df=load_orders(); s=request.args.get("search",""); sf=request.args.get("status","all")
    if s: df=df[df["CustomerName"].str.contains(s,case=False,na=False)|df["OrderID"].astype(str).str.contains(s,case=False,na=False)]
    if sf!="all": df=df[df["Status"].str.lower()==sf.lower()]
    df2=df.copy(); df2["Date"]=df2["Date"].dt.strftime("%Y-%m-%d")
    return render_template("orders.html",orders=df2.to_dict(orient="records"),search=s,status_filter=sf)

@app.route("/customers")
@login_required
def customers():
    df=load_orders(); s=request.args.get("search","")
    if s: df=df[df["CustomerName"].str.contains(s,case=False,na=False)|df["Phone"].astype(str).str.contains(s,case=False,na=False)]
    cdf=df.groupby(["CustomerName","Phone"]).agg(total_orders=("OrderID","count"),total_spent=("Amount","sum"),last_order=("Date","max")).reset_index()
    cdf["last_order"]=cdf["last_order"].dt.strftime("%Y-%m-%d")
    return render_template("customers.html",customers=cdf.to_dict(orient="records"),search=s)

@app.route("/pending")
@login_required
def pending():
    df=load_orders(); p=df[df["Status"]=="Pending"].copy(); p["Date"]=p["Date"].dt.strftime("%Y-%m-%d")
    return render_template("pending.html",orders=p.to_dict(orient="records"))

@app.route("/whatsapp")
@login_required
def whatsapp():
    df=load_orders()
    pend=df[df["Status"]=="Pending"].copy(); sent=df[df["Status"].isin(["Sent","Delivered"])].copy()
    pend["Date"]=pend["Date"].dt.strftime("%Y-%m-%d"); sent["Date"]=sent["Date"].dt.strftime("%Y-%m-%d")
    return render_template("whatsapp.html",pending=pend.to_dict(orient="records"),sent=sent.to_dict(orient="records"))

@app.route("/delivered")
@login_required
def delivered():
    df=load_orders(); d=df[df["Status"]=="Delivered"].copy(); d["Date"]=d["Date"].dt.strftime("%Y-%m-%d")
    return render_template("delivered.html",orders=d.to_dict(orient="records"))

@app.route("/reports")
@login_required
def reports():
    df=load_orders(); charts={}
    if not df.empty:
        for name,fn in [("revenue",chart_revenue),("status",chart_donut),("products",chart_products),("customers",chart_customers)]:
            try: charts[name]=fn(df)
            except: pass
    df2=df.dropna(subset=["Date"]).copy(); df2["Month"]=df2["Date"].dt.strftime("%b %Y")
    monthly=df2.groupby("Month").agg(orders=("OrderID","count"),revenue=("Amount","sum")).reset_index().to_dict(orient="records")
    return render_template("reports.html",stats=get_stats(df),charts=charts,insights=ai_insights(df),monthly=monthly)

@app.route("/settings")
@login_required
def settings():
    return render_template("settings.html")

# ─── CRUD ROUTES ─────────────────────────────────────────
@app.route("/add",methods=["POST"])
@login_required
def add_order():
    df=load_orders()
    new_id=f"ORD-{len(df)+1:03d}"
    new_row=pd.DataFrame([{"OrderID":new_id,"CustomerName":request.form.get("name"),
        "Phone":request.form.get("phone"),"Product":request.form.get("product"),
        "Amount":float(request.form.get("amount",0)),"Status":"Pending","Date":datetime.now()}])
    save_orders(pd.concat([df,new_row],ignore_index=True))
    flash("✅ Order added successfully!","success")
    return redirect(url_for("orders"))

@app.route("/send/<order_id>")
@login_required
def send_whatsapp(order_id):
    df=load_orders(); order=df[df["OrderID"]==order_id]
    if order.empty: return jsonify({"success":False,"error":"Order not found"}),404
    o=order.iloc[0]; phone=fmt_phone(o["Phone"])
    msg=(f"Assalam o Alaikum {o['CustomerName']}! 🎉\nAapka order confirm ho gaya!\n\n"
         f"📦 Product: *{o['Product']}*\n💰 Total: *Rs {o['Amount']:,.0f}*\n🆔 Order ID: *{o['OrderID']}*\n\nShukriya! 🙏")
    try:
        pywhatkit.sendwhatmsg_instantly(phone,msg,wait_time=10,tab_close=True)
        df.loc[df["OrderID"]==order_id,"Status"]="Sent"; save_orders(df)
        return jsonify({"success":True,"message":"WhatsApp sent!"})
    except Exception as e:
        return jsonify({"success":False,"error":str(e)}),500

@app.route("/send_bulk",methods=["POST"])
@login_required
def send_bulk():
    df=load_orders(); results=[]
    for _,o in df[df["Status"]=="Pending"].iterrows():
        phone=fmt_phone(o["Phone"])
        msg=f"Assalam o Alaikum {o['CustomerName']}! Order *{o['Product']}* confirm — Rs {o['Amount']:,.0f} | ID: {o['OrderID']} 🎉"
        try:
            pywhatkit.sendwhatmsg_instantly(phone,msg,wait_time=10,tab_close=True)
            df.loc[df["OrderID"]==o["OrderID"],"Status"]="Sent"
            results.append({"id":o["OrderID"],"success":True})
        except Exception as e:
            results.append({"id":o["OrderID"],"success":False,"error":str(e)})
    save_orders(df); return jsonify({"results":results})

@app.route("/delete/<order_id>",methods=["POST"])
@login_required
def delete_order(order_id):
    df=load_orders(); save_orders(df[df["OrderID"]!=order_id])
    flash("🗑️ Order deleted.","info")
    return redirect(request.referrer or url_for("orders"))

@app.route("/update_status/<order_id>",methods=["POST"])
@login_required
def update_status(order_id):
    df=load_orders()
    df.loc[df["OrderID"]==order_id,"Status"]=request.form.get("status","Pending")
    save_orders(df)
    return redirect(request.referrer or url_for("orders"))

@app.route("/import_excel",methods=["POST"])
@login_required
def import_excel():
    f=request.files.get("file")
    if not f: return jsonify({"success":False,"error":"No file"}),400
    try:
        new_df=pd.read_excel(f); existing=load_orders()
        merged=pd.concat([existing,new_df],ignore_index=True).drop_duplicates(subset=["OrderID"])
        save_orders(merged); return jsonify({"success":True,"imported":len(new_df)})
    except Exception as e:
        return jsonify({"success":False,"error":str(e)}),500

# ═══════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("\n╔══════════════════════════════════════════╗")
    print("║      OrderPro v3.0  —  Starting...       ║")
    print("╚══════════════════════════════════════════╝")
    if os.path.exists(EXCEL_FILE):
        df=load_orders()
        print(f"✅  {len(df)} orders loaded | Revenue: Rs {df['Amount'].sum():,.0f}")
    else:
        print("⚠️  orders.xlsx missing — run: python setup.py")
    print("🔐  Login: admin / admin123")
    print("🌐  Opening: http://127.0.0.1:5000\n")
    threading.Timer(1.2,lambda: webbrowser.open("http://127.0.0.1:5000")).start()
    app.run(debug=True,host="0.0.0.0",port=5000,use_reloader=False)